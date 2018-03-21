'''
import pygeohash as pgh

pgh.encode(42.6, -5.6)
# >>> 'ezs42e44yx96'

pgh.encode(42.6, -5.6, precision=5)
# >>> 'ezs42'

pgh.decode('ezs42')
# >>> ('42.6', '-5.6')

pgh.geohash_approximate_distance('bcd3u', 'bc83n')
# >>> 625441
'''

import pygeohash as pgh
from openpyxl import Workbook
from openpyxl import load_workbook


# step1: 载入工作表
workbook = load_workbook('source_excel.xlsx')
sheet = workbook.get_sheet_by_name("Sheet1")


# 指定表格数据读取范围
start_loc_range = sheet['A2':'A20']
end_loc_range = sheet['B2':'B20']
# 读取数据
start_loc_geohash = []
end_loc_geohash = []
# 读取start_loc列
for row_of_cell in start_loc_range:
    for cell in row_of_cell:
        start_loc_geohash.append(cell.value)
# 读取end_loc列
for row_of_cell in end_loc_range:
    for cell in row_of_cell:
        end_loc_geohash.append(cell.value)


# step2: 转换为坐标
start_loc_coord = []
end_loc_coord = []
for i in range(len(start_loc_geohash)):
    decode = pgh.decode(start_loc_geohash[i])
    start_loc_coord.append(str(decode))
for i in range(len(end_loc_geohash)):
    decode = pgh.decode(end_loc_geohash[i])
    end_loc_coord.append(str(decode))


# step3: 保存坐标xlsx文件
coord_book = Workbook()
coord_sheet = coord_book.get_sheet_by_name("Sheet")
coord_book.save('coords.xlsx')
coord_sheet['A1'] = 'start_loc'
coord_sheet['B1'] = 'end_loc'

for i in range(len(start_loc_coord)):
    coord_sheet.cell(row=i+2, column=1).value = start_loc_coord[i]
for i in range(len(end_loc_coord)):
    coord_sheet.cell(row=i+2, column=2).value = end_loc_coord[i]

coord_book.save('coords.xlsx')